from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

FONT = "Arial"
INK = "1F2933"
HEADER_FILL = PatternFill("solid", fgColor="1F2933")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BASE_FONT = Font(name=FONT, size=10, color=INK)
TITLE_FONT = Font(name=FONT, bold=True, size=16, color=INK)
SUB_FONT = Font(name=FONT, size=10, color="6B7280")
BOLD = Font(name=FONT, bold=True, size=10, color=INK)
BAR_FILL = PatternFill("solid", fgColor="0066FF")
BAR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
GROUP_FILL = PatternFill("solid", fgColor="EEF2FF")
thin = Side(style="thin", color="E4E4E7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
GREEN = PatternFill("solid", fgColor="DCFCE7")
AMBER = PatternFill("solid", fgColor="FEF3C7")
RED = PatternFill("solid", fgColor="FEE2E2")
RISK_FONT = {"At risk": "991B1B", "Watch": "92400E", "On track": "166534"}
RISK_FILLP = {"At risk": RED, "Watch": AMBER, "On track": GREEN}

MENTEES = [
    ["Aisha Khan","aisha.k@email.com","Phoenix Clan","Intermediate","Toronto, CA","6/12","Backend Foundations","3. Add authentication",72,88,84,"Up","Positive","On track",1,"2h ago",0,"May 15","Thu, May 29",85,78,92,70,"Two submissions waiting. JWT refresh-token blocker flagged early. Pair on auth Thursday."],
    ["Diego Morales","diego.m@email.com","Phoenix Clan","Intermediate","Lahore, PK","6/12","Frontend Craft","1. Component composition",41,79,58,"Up","Neutral","Watch",2,"1d ago",0,"May 18","-",55,72,95,60,"Power outages and a shared machine. Struggling despite effort. Protect deadlines, shift due times to AM PKT."],
    ["Priya Nair","priya.n@email.com","Phoenix Clan","Intermediate","Bengaluru, IN","6/12","Not started","-",31,34,44,"Down","Low","At risk",0,"6d ago",2,"May 20","Wed, May 28",40,38,50,65,"Disengagement risk. No login for 6 days, two nudges unanswered. Direct warm check-in. Psychologist session held May 20."],
    ["Liam Walsh","liam.w@email.com","Phoenix Clan","Intermediate","Dublin, IE","6/12","Not started","-",95,76,97,"Flat","Positive","On track",0,"5h ago",0,"-","-",96,65,70,90,"Top of cohort on output, almost always early. May be coasting on easy ground. Add a stretch track."],
    ["Fatima Noor","fatima.n@email.com","Phoenix Clan","Intermediate","Karachi, PK","6/12","Not started","-",64,71,79,"Up","Positive","On track",1,"3h ago",0,"-","-",80,85,75,72,"Solid and improving three weeks running. One async/await blocker, working at it."],
    ["Tomas Berg","tomas.b@email.com","Phoenix Clan","Intermediate","Oslo, NO","6/12","Not started","-",53,49,62,"Down","Neutral","Watch",1,"2d ago",0,"-","-",60,70,64,68,"Drifting, two late tasks this week. Still responsive. A short check-in now prevents a bigger slide."],
]
# week-1 example, in line with the schedule.
# journaling/5, reading/5, mindset/5, eng/5, dean/5, core assigned, core done,
# core %, grind hours /10, family Yes/No, points /10, risk, feedback
WEEK1 = {
    "Aisha Khan": (5, 4, 5, 5, 4, "Add authentication (JWT)", "Auth started, refresh-token flow next", 70, 8, "Yes", 8, "On track", "Strong week. Pair on auth Thursday."),
    "Diego Morales": (4, 3, 4, 4, 4, "Component composition", "Reading done, dashboard in late", 60, 7, "Yes", 6, "Watch", "Good effort against real constraints. Protect deadlines."),
    "Priya Nair": (1, 0, 1, 1, 1, "API integration exercise", "Little progress, re-engaging", 20, 2, "Yes", 3, "At risk", "Warm check-in done. Lighter load next week, small wins first."),
    "Liam Walsh": (5, 5, 5, 5, 5, "Auth middleware + rate-limit stretch", "Middleware done early", 100, 9, "Yes", 9, "On track", "Way ahead. Give him a real stretch."),
    "Fatima Noor": (5, 4, 5, 5, 4, "Async/await refactor", "In progress, event-loop reading done", 75, 8, "Yes", 8, "On track", "Climbing nicely. Unblock async/await."),
    "Tomas Berg": (4, 3, 4, 3, 3, "Form validation library", "Submitted late, regex quiz pending", 55, 6, "Yes", 5, "Watch", "Slipping a little. Short check-in this week."),
}

WEEKS = 12
START = date(2026, 6, 1)
START = START - timedelta(days=START.weekday())


def style_header(ws, headers):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def dv(ws, col, options, last, src=False):
    f = options if src else '"' + ",".join(options) + '"'
    d = DataValidation(type="list", formula1=f, allow_blank=True)
    d.add(f"{col}2:{col}{last}")
    ws.add_data_validation(d)


wb = Workbook()
MENTEE_LIST = "'Mentee Tracker'!$A$2:$A$60"

# ---------------- Guide ----------------
g = wb.active
g.title = "Guide"
g.sheet_view.showGridLines = False
g["A1"] = "Dev Weekend - Mentee Tracker"
g["A1"].font = TITLE_FONT
g["A2"] = "A simple shared sheet for mentors to track each mentee while Pathment is being built."
g["A2"].font = SUB_FONT
rows = [
    ("", ""),
    ("How to use", ""),
    ("1.", "Open in Google Sheets (File, Import, Upload) or Excel. Share with your co-mentors."),
    ("2.", "Weekly Tracker is what you fill each week, in line with the schedule. For each mentee, mark how many of the five weekday days they did each habit and talk, what core work was assigned and how much got done, the weekend grind hours, and give points out of 10."),
    ("3.", "Mentee Tracker is the current snapshot, one row per mentee, with their points to date."),
    ("4.", "Daily Log is the day-by-day version of the same schedule. 1-on-1 Log is for sessions."),
    ("5.", "Keep it honest. The sheet is to help people, not to grade them."),
    ("", ""),
    ("The schedule it follows", ""),
    ("", "Weekdays run journaling, morning reading, the Mindset talk at breakfast, core work, the Engineering talk at lunch, and the Dean talk at dinner. Weekends are the ten-hour grind plus family time. The weekly columns line up with exactly that."),
    ("", ""),
    ("What was assigned, done, and how much", ""),
    ("", "For the habits and talks, the target is the five weekday days, so 4 out of 5 means four days done. For core work you write what was assigned, what they completed, and the percent done. For the grind, hours out of ten. That is the assigned, completed, and how-much, in line with the schedule."),
    ("", ""),
    ("Built to import into Pathment later", ""),
    ("", "Keep the headers and dropdown values as they are, one mentee per row. Email is the key that links a mentee across tabs. See import-format.md."),
]
r = 4
for a, b in rows:
    g[f"A{r}"], g[f"B{r}"] = a, b
    g[f"A{r}"].font = BASE_FONT
    g[f"B{r}"].font = BASE_FONT
    g[f"B{r}"].alignment = WRAP
    r += 1
for cell in ("A5", "A12", "A15", "A18"):
    g[cell].font = Font(name=FONT, bold=True, size=11, color=INK)
sr = r + 1
g[f"A{sr}"] = "At a glance (updates from the tracker)"
g[f"A{sr}"].font = Font(name=FONT, bold=True, size=11, color=INK)
for i, (label, formula) in enumerate([
    ("Mentees tracked", "=COUNTA('Mentee Tracker'!A2:A200)"),
    ("On track", "=COUNTIF('Mentee Tracker'!N2:N200,\"On track\")"),
    ("Watch", "=COUNTIF('Mentee Tracker'!N2:N200,\"Watch\")"),
    ("At risk", "=COUNTIF('Mentee Tracker'!N2:N200,\"At risk\")"),
    ("Average on-time %", "=IFERROR(ROUND(AVERAGE('Mentee Tracker'!K2:K200),0),0)"),
    ("Average points (filled weeks)", "=IFERROR(ROUND(AVERAGEIF('Weekly Tracker'!N2:N2000,\">0\"),1),0)"),
]):
    g[f"A{sr + 1 + i}"] = label
    g[f"A{sr + 1 + i}"].font = BASE_FONT
    g[f"B{sr + 1 + i}"] = formula
    g[f"B{sr + 1 + i}"].font = BOLD
g.column_dimensions["A"].width = 30
g.column_dimensions["B"].width = 100

# ---------------- Weekly Tracker (schedule-aligned marking sheet) ----------------
w = wb.create_sheet("Weekly Tracker")
w.sheet_view.showGridLines = False
wheaders = ["Week of", "Mentee", "Email",
            "Journaling /5", "Reading /5", "Mindset talk /5", "Eng. talk /5", "Dean talk /5",
            "Core work assigned", "Core work completed", "Core %",
            "Grind h /10", "Family time", "Points /10", "Risk", "Feedback / focus for next week"]
style_header(w, wheaders)
ncol = len(wheaders)
wrap_cols = {9, 10, 16}
center_cols = {4, 5, 6, 7, 8, 11, 12, 13, 14, 15}
row = 2
for wk in range(WEEKS):
    d0 = START + timedelta(weeks=wk)
    bar = f"Week of {d0.strftime('%b')} {d0.day}, {d0.year}"
    w.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncol)
    bc = w.cell(row=row, column=1, value=bar)
    bc.fill = BAR_FILL
    bc.font = BAR_FONT
    bc.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    w.row_dimensions[row].height = 22
    row += 1
    for m in MENTEES:
        name, email = m[0], m[1]
        if wk == 0:
            ex = WEEK1[name]
            vals = [d0, name, email] + list(ex)
        else:
            vals = [d0, name, email] + [""] * (ncol - 3)
        for c, v in enumerate(vals, start=1):
            cell = w.cell(row=row, column=c, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            if c == 1:
                cell.number_format = "mmm d, yyyy"
                cell.alignment = TOP
            elif c in wrap_cols:
                cell.alignment = WRAP
            elif c in center_cols:
                cell.alignment = CENTER
            else:
                cell.alignment = TOP
        row += 1
last = row - 1
for i, wd in enumerate([13, 15, 20, 8, 8, 8, 8, 8, 30, 30, 7, 8, 8, 8, 11, 36], start=1):
    w.column_dimensions[get_column_letter(i)].width = wd
w.row_dimensions[1].height = 30
w.freeze_panes = "B2"
for col in ("D", "E", "F", "G", "H"):
    dv(w, col, [str(n) for n in range(0, 6)], last)
dv(w, "L", [str(n) for n in range(0, 11)], last)
dv(w, "M", ["Yes", "No"], last)
dv(w, "N", [str(n) for n in range(0, 11)], last)
dv(w, "O", ["On track", "Watch", "At risk"], last)
rng = f"O2:O{last}"
for val in ("At risk", "Watch", "On track"):
    w.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{val}"'], fill=RISK_FILLP[val], font=Font(name=FONT, size=10, color=RISK_FONT[val])))

# ---------------- Mentee Tracker ----------------
t = wb.create_sheet("Mentee Tracker")
t.sheet_view.showGridLines = False
headers = ["Mentee", "Email", "Clan", "Level", "Location", "Week", "Current roadmap", "Current step",
           "Absolute %", "Relative %", "On-time %", "Momentum", "Sentiment", "Risk",
           "Open blockers", "Last active", "Nudges (wk)", "Last 1:1", "Next 1:1",
           "Consistency", "Communication", "Resilience", "Independence", "Mentor notes / next steps",
           "Points to date"]
style_header(t, headers)
for m in MENTEES:
    t.append(m)
for _ in range(12):
    t.append([""] * (len(headers) - 1))
ncols = len(headers)
nrows = 1 + len(MENTEES) + 12
center = set(range(9, 24)) | {25}
for rr in range(2, nrows + 1):
    t.cell(row=rr, column=25).value = (
        f"=IF(B{rr}=\"\",\"\",SUMIF('Weekly Tracker'!$C:$C,B{rr},'Weekly Tracker'!$N:$N))"
    )
    for c in range(1, ncols + 1):
        cell = t.cell(row=rr, column=c)
        cell.font = BOLD if c == 25 else BASE_FONT
        cell.border = BORDER
        cell.alignment = WRAP if c == 24 else (CENTER if c in center else TOP)
widths = [16,22,14,13,15,7,18,22,10,10,10,11,11,11,8,11,9,10,12,11,13,10,12,52,13]
for i, wd in enumerate(widths, start=1):
    t.column_dimensions[get_column_letter(i)].width = wd
t.row_dimensions[1].height = 28
t.freeze_panes = "C2"
dv(t, "L", ["Up", "Flat", "Down"], nrows)
dv(t, "M", ["Positive", "Neutral", "Low"], nrows)
dv(t, "N", ["On track", "Watch", "At risk"], nrows)
rng = f"N2:N{nrows}"
for val in ("At risk", "Watch", "On track"):
    t.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=[f'"{val}"'], fill=RISK_FILLP[val], font=Font(name=FONT, size=10, color=RISK_FONT[val])))

# ---------------- Daily Log ----------------
d = wb.create_sheet("Daily Log")
d.sheet_view.showGridLines = False
dheaders = ["Date","Mentee","Journaling","Morning reading","Mindset talk (breakfast)","Core work",
            "Engineering talk (lunch)","Dean talk (dinner)","Weekend grind","Family time","Tasks done today","Note"]
style_header(d, dheaders)
daily = [
    ["Mon May 26","Aisha Khan","Yes","Yes","Yes","Yes","Yes","Yes","-","-","SQL Fundamentals quiz","Mindset talk hit home, kept focus through the SQL quiz."],
    ["Tue May 27","Aisha Khan","Yes","Yes","Yes","Yes","Yes","No","-","-","","Missed the dean talk, ran late on the API work."],
    ["Sat May 31","Aisha Khan","-","-","-","-","-","-","Yes","Yes","Auth roadmap progress","Ten hour grind on auth, then family dinner."],
]
for rowd in daily:
    d.append(rowd)
for _ in range(24):
    d.append([""] * len(dheaders))
dn = 1 + len(daily) + 24
for rr in range(2, dn + 1):
    for c in range(1, len(dheaders) + 1):
        cell = d.cell(row=rr, column=c)
        cell.font = BASE_FONT
        cell.border = BORDER
        cell.alignment = WRAP if c == len(dheaders) else (CENTER if 3 <= c <= 10 else TOP)
for i, wd in enumerate([12,16,11,15,22,11,20,18,13,12,20,46], start=1):
    d.column_dimensions[get_column_letter(i)].width = wd
d.row_dimensions[1].height = 40
d.freeze_panes = "C2"
dv(d, "B", MENTEE_LIST, dn, src=True)
yesno = DataValidation(type="list", formula1='"Yes,No,-"', allow_blank=True)
yesno.add(f"C2:J{dn}")
d.add_data_validation(yesno)

# ---------------- 1-on-1 Log ----------------
o = wb.create_sheet("1-on-1 Log")
o.sheet_view.showGridLines = False
oheaders = ["Date","Mentee","Logged by","Type","Sentiment","Summary","Issues raised","Next steps"]
style_header(o, oheaders)
ones = [
    ["May 15","Aisha Khan","Sarah Chen","1:1","Positive","Discussed career goals, wants to move into backend. Prefers async written feedback.","","Share backend track. Pair on auth Thursday."],
    ["May 20","Priya Nair","Dr. Maya Brooks","Psychological session","Neutral","Initial session, anxiety around pace and comparison to peers. Coping strategies discussed.","Anxiety, pace","Follow-up in two weeks."],
]
for rowo in ones:
    o.append(rowo)
for _ in range(24):
    o.append([""] * len(oheaders))
on = 1 + len(ones) + 24
for rr in range(2, on + 1):
    for c in range(1, len(oheaders) + 1):
        cell = o.cell(row=rr, column=c)
        cell.font = BASE_FONT
        cell.border = BORDER
        cell.alignment = WRAP if c in (6, 7, 8) else (CENTER if c == 5 else TOP)
for i, wd in enumerate([10,16,16,22,11,46,28,40], start=1):
    o.column_dimensions[get_column_letter(i)].width = wd
o.freeze_panes = "A2"
dv(o, "B", MENTEE_LIST, on, src=True)
dv(o, "D", ["1:1", "Pairing session", "Career chat", "Psychological session", "Check-in"], on)
dv(o, "E", ["Positive", "Neutral", "Low"], on)

out = r"D:\Projects\pathment\mentor-tracking-sheet\Dev Weekend - Mentee Tracker.xlsx"
wb.save(out)
print("saved", out)
